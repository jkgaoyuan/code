import openpyxl
# 创建 excel 
# create_wb = openpyxl.Workbook("test1.xlsx")  
# 加载excel 文件
read_wb = openpyxl.load_workbook('E:\pycharm\\code\\AI_AGENT\\test.xlsx')
# 激活对象
active_wb = read_wb.active
print(active_wb)
maxrow = active_wb.max_row  
print('这是最大行%d' %maxrow)
print('这是最大列%d' %maxclo)
columns = active_wb['A1':'A%d' %maxrow ]
print(active_wb['A2'].value) # 使用value方法可以读取 数值 否则将直接输出对应的坐标
for i in columns:
    print(i)
    for cells in i:
        print(cells.value)


# read_wb.read_only